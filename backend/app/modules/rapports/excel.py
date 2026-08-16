"""Génération des rapports Excel (openpyxl). Contrairement à WeasyPrint,
openpyxl est une dépendance pure Python, importée normalement en tête de
module — aucune bibliothèque native, aucune raison de la différer."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.rapports.pdf import Colonne


def _nouvelle_feuille(classeur: Workbook, titre: str) -> Worksheet:
    feuille = classeur.active
    assert feuille is not None  # toujours vrai pour un Workbook() fraîchement créé
    feuille.title = titre[:31] or "Rapport"  # limite Excel : 31 caractères
    return feuille


def excel_tableau(
    *,
    titre: str,
    colonnes: list[Colonne],
    lignes: list[dict[str, object]],
    totaux: list[dict[str, object]] | None = None,
) -> bytes:
    classeur = Workbook()
    feuille = _nouvelle_feuille(classeur, titre)

    feuille.append([colonne.libelle for colonne in colonnes])
    for cellule in feuille[1]:
        cellule.font = Font(bold=True)

    for ligne in lignes:
        feuille.append([ligne.get(colonne.cle, "") for colonne in colonnes])

    if totaux:
        feuille.append([cellule.get("valeur", "") for cellule in totaux])
        for cellule in feuille[feuille.max_row]:
            cellule.font = Font(bold=True)

    for indice, colonne in enumerate(colonnes, start=1):
        feuille.column_dimensions[get_column_letter(indice)].width = max(
            12, len(colonne.libelle) + 2
        )

    tampon = BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


def excel_cle_valeur(*, titre: str, paires: list[tuple[str, str]]) -> bytes:
    classeur = Workbook()
    feuille = _nouvelle_feuille(classeur, titre)

    feuille.append(["Indicateur", "Valeur"])
    for cellule in feuille[1]:
        cellule.font = Font(bold=True)

    for libelle, valeur in paires:
        feuille.append([libelle, valeur])

    feuille.column_dimensions["A"].width = 35
    feuille.column_dimensions["B"].width = 20

    tampon = BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()
